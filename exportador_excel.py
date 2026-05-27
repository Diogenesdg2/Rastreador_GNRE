import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


def gerar_excel(dados, caminho_saida):
    """
    Gera a planilha Excel com os dados extraídos,
    incluindo totalizador de Total e Valor Pago.
    """

    if not dados:
        dados = []

    df = pd.DataFrame(dados)

    colunas = [
        "tipo",
        "data_vencimento",
        "documento_origem",
        "total_recolher",
        "codigo_barras",
        "valor_pago",
        "data_pagamento",
        "hora_pagamento",
        "arquivo",
    ]

    for coluna in colunas:
        if coluna not in df.columns:
            df[coluna] = ""

    df = df[colunas]

    # Garante que documento_origem e codigo_barras sejam texto,
    # para o Excel não remover zeros à esquerda.
    df["documento_origem"] = df["documento_origem"].astype(str)
    df["codigo_barras"] = df["codigo_barras"].astype(str)

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="GNRE")

        ws = writer.book["GNRE"]

        # Formatar como texto
        coluna_documento = colunas.index("documento_origem") + 1
        coluna_codigo = colunas.index("codigo_barras") + 1

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=coluna_documento).number_format = "@"
            ws.cell(row=row, column=coluna_codigo).number_format = "@"

        # Colunas para totalização
        coluna_total = colunas.index("total_recolher") + 1
        coluna_valor_pago = colunas.index("valor_pago") + 1

        linha_total = ws.max_row + 1

        # Rótulo do totalizador
        ws.cell(row=linha_total, column=1).value = "TOTAL GERAL"

        # Fórmulas de soma
        ws.cell(row=linha_total, column=coluna_total).value = (
            f"=SUM({ws.cell(row=2, column=coluna_total).coordinate}:"
            f"{ws.cell(row=linha_total - 1, column=coluna_total).coordinate})"
        )

        ws.cell(row=linha_total, column=coluna_valor_pago).value = (
            f"=SUM({ws.cell(row=2, column=coluna_valor_pago).coordinate}:"
            f"{ws.cell(row=linha_total - 1, column=coluna_valor_pago).coordinate})"
        )

        # Formatação da linha de total
        fill_total = PatternFill(
            start_color="D9EAD3",
            end_color="D9EAD3",
            fill_type="solid"
        )

        fonte_total = Font(bold=True)

        for col in range(1, len(colunas) + 1):
            celula = ws.cell(row=linha_total, column=col)
            celula.font = fonte_total
            celula.fill = fill_total
            celula.alignment = Alignment(horizontal="center")

        # Formatação numérica das colunas de valores
        for row in range(2, linha_total + 1):
            ws.cell(row=row, column=coluna_total).number_format = '#,##0.00'
            ws.cell(row=row, column=coluna_valor_pago).number_format = '#,##0.00'

        # Largura das colunas
        larguras = {
            "A": 24,
            "B": 18,
            "C": 22,
            "D": 18,
            "E": 60,
            "F": 18,
            "G": 18,
            "H": 18,
            "I": 80,
        }

        for coluna, largura in larguras.items():
            ws.column_dimensions[coluna].width = largura


# Compatibilidade com o main.py
exportar_excel = gerar_excel